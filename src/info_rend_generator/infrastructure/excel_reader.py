from __future__ import annotations

import warnings
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from info_rend_generator.domain.money import quantize_money
from info_rend_generator.domain.transactions import Transaction

warnings.filterwarnings(
    "ignore",
    message="(Data Validation|Slicer List) extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl",
)


def read_excel_transactions(path: str | Path) -> list[Transaction]:
    workbook = _load_workbook(path)
    worksheet = _transaction_worksheet(workbook)
    headers = _headers(worksheet)

    date_column = _required_column(headers, "data lancamento")
    description_column = _required_column(headers, "descricao")
    amount_column = _required_column(headers, "valor")
    type_column = _optional_column(headers, "tipotransacao")
    classification_column = _optional_column(headers, "classificacaocontabil")

    transactions: list[Transaction] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        date_value = _cell(row, date_column)
        amount_value = _cell(row, amount_column)
        if date_value in (None, "") or amount_value in (None, ""):
            continue

        description = str(_cell(row, description_column) or "")
        classification = str(_cell(row, classification_column) or "")
        transaction_type = str(_cell(row, type_column) or "").upper()

        transactions.append(
            Transaction(
                date=_parse_date(date_value),
                trntype=transaction_type,
                amount=quantize_money(amount_value),
                name=description,
                memo=classification,
            )
        )

    return transactions


def read_excel_raw_data(path: str | Path) -> list[dict[str, Any]]:
    workbook = _load_workbook(path)
    rows: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        sheet_headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            for column_index, value in enumerate(row, start=1):
                header = (
                    sheet_headers[column_index - 1] if column_index <= len(sheet_headers) else None
                )
                rows.append(
                    {
                        "sheet": worksheet.title,
                        "row": row_index,
                        "column": column_index,
                        "header": header,
                        "value": _raw_value(value),
                    }
                )

    return rows


def _load_workbook(path: str | Path):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Excel input requires openpyxl. Install dependencies with "
            "`python -m pip install -e .`."
        ) from exc

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            module="openpyxl",
        )
        return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _transaction_worksheet(workbook):
    for worksheet in workbook.worksheets:
        normalized_headers = set(_headers(worksheet))
        required = {"data lancamento", "descricao", "valor"}
        if required.issubset(normalized_headers):
            return worksheet
    raise ValueError(
        "Could not find an Excel sheet with the required columns: "
        "Data Lançamento, Descrição, Valor."
    )


def _headers(worksheet) -> dict[str, int]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {
        _normalize_header(value): index
        for index, value in enumerate(first_row)
        if value not in (None, "")
    }


def _required_column(headers: dict[str, int], name: str) -> int:
    column = _optional_column(headers, name)
    if column is None:
        raise ValueError(f"Missing required Excel column: {name}")
    return column


def _optional_column(headers: dict[str, int], name: str) -> int | None:
    return headers.get(name)


def _cell(row: tuple[Any, ...], column: int | None) -> Any:
    if column is None or column >= len(row):
        return None
    return row[column]


def _parse_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))


def _normalize_header(value: Any) -> str:
    normalized = str(value).strip().lower()
    replacements = {
        "ç": "c",
        "ã": "a",
        "á": "a",
        "à": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return normalized


def _raw_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value
